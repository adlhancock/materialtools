# -*- coding: utf-8 -*-
""" tools for exporting materialproperty data
Created on Wed Nov  9 14:11:37 2016

@author: dhancock
"""
import materialtools
def textfile(materialdata,filename,verbose=False):
    print('exporting text file not yet supported')
    pass

def json(materialdata,filename,verbose=False):
        '''
        exports material property data from MatML_Data object in json format
        '''
        data = dict(materialdata)
        from json import dump
        with open(filename,'w') as f:
            dump(data,
                 f,
                 indent=4)
        if verbose is True: print('exported json to',filename)
        return
        
def csv(materialdata,filename,verbose=False):
    """ exports materialdata as a csv file
    
    notes:
        produces a somewhat indented tree-type structure for clarity
    """
    def tabulate_csv_data(outfile,materialdata):
        from materialtools import MaterialProperty, MaterialParameter
        for material in materialdata.values():
            print("{},{}".format("Material Name",material['Name']), 
                  file = outfile)
            for item in [i for i in material 
                        if type(material[i]) is not MaterialProperty]:
                print(",{},{}".format(item,material[item]),file = outfile)
                
            for prop in [p for p in material.values() 
                                        if type(p) is MaterialProperty]:
                print("{},{}".format(",Property Name",prop["Name"]), 
                      file = outfile)
                """
                for item in [i for i in prop if type(prop[i]) is not MaterialParameter]:
                    print(",,{},{}".format(item,prop[item]), file = outfile)
                """
                for parm in [p for p in prop.values() 
                                        if type(p) is MaterialParameter]:
                    print("{},{}".format(",,Parameter Name",parm["Name"]), 
                          file = outfile)
                    print(",,,Parameter Values", 
                          file = outfile, 
                          end=",")
                    [print(v, file=outfile, end=",") for v in parm["Values"]]
                    print('', 
                          file = outfile)
                    print(",,,Parameter Units", 
                          file = outfile, 
                          end=",")
                    [print(u, file=outfile, end=",") for u in parm["Units"]] 
                    print('', 
                          file = outfile)
                print("", file = outfile)
                
            print("", file = outfile)
        return
        
    try:
        with open(filename, mode = 'x') as outfile:
            print('creating new file: {}'.format(filename))
            print(file = outfile)
    except FileExistsError:
        print('overwriting {}'.format(filename))
    finally:
        with open(filename, mode = 'w') as outfile:
            tabulate_csv_data(outfile,materialdata)
    return
    
def xlsx(materialdata,filepath,verbose=False):
    """ writes to an excel file"""
    
    from openpyxl import Workbook
    from materialtools import Material, MaterialProperty, MaterialParameter
    import os
    
    ## create directory if necessary
    if filepath.endswith('/'): filepath = filepath[:-1]
    try: os.stat(filepath)
    except FileNotFoundError: 
        os.mkdir(filepath) 
        print("creating directory:",filepath)
    except: raise
        
    ## cycle through materials
    materials = [materialdata[m] for m in materialdata if type(materialdata[m]) is Material]
    for m in materials:
        ## create empty workbook and dictionary of worksheets
        wb = Workbook()
        
        ## create materialdata metadata worksheet
        ws = wb.active
        ws.title = "Metadata"

        ## this bit just takes the "useful" bits of metadata
        ws.append(["MaterialName",m.name])
        ws.append(["Condition",m["Condition"]])
        try: ws.append(["Description",m["Description"]])
        except: pass
        ws.append(["DataSource"]+[materialdata.source])
        ws.append(["Filename"]+materialdata.filename)
        
        
        ## get the property names, abbreviate, and add a table to the front ws.
        properties = [m[p] for p in m if type(m[p]) is MaterialProperty]
        propertynames = [p['PropertyName'] for p in properties]
        abbreviate = lambda name: ''.join([w[0] for w in name.split(' ')])
        abbreviations = [abbreviate(p) for p in propertynames]
        ws.append(["Propertynames"]+propertynames)        
        ws.append(["Abbreviations"]+abbreviations)
        """
        for row in zip(abbreviations,properties):
            ws.append([row[0],row[1]["Name"]])
        """ 
        ## cycle through properties
        for prop in properties:
            propertyname = prop["PropertyName"]
            
            ## create a worksheet for each property 
            ws=wb.create_sheet(title=abbreviate(propertyname))
            
            ## cycle through parameters for each property
            parameters = [prop[par] for par in prop 
                                if type(prop[par]) is MaterialParameter]
            #print("XXX",parameters,"XXX")                                                          #### debug
            parameternames = [par["ParameterName"] for par in parameters]

            if verbose is True: 
                print('\t',prop["PropertyName"],':',parameternames)

            ws.append(parameternames)
            units = [par["Units"][0] for par in parameters]
            ws.append(units)
            for i,par in enumerate(parameters):
                for j,val in enumerate(par["Values"]):
                    ws.cell(row=j+3,column=i+1,value=val)
                    
            ## if master property values not already given, add them if they are there
            if prop["PropertyName"] not in parameternames:
                if "Values" in prop:
                    ws.cell(row = 1, column = len(parameters)+1,value=prop["PropertyName"])
                    ws.cell(row = 2, column = len(parameters)+1,value=prop["Units"][0])
                    for j, val in enumerate(prop["Values"]):
                        ws.cell(row = j+3, column=len(parameters)+1,value=val)
                    
            ## add columns for other fields
            pass
        ## save workbook for each material
        filename = '{}/{}-{}.xlsx'.format(filepath,m.name,materialdata.source)
        tryagain = True
        writefile = True
        while tryagain is True:
            try: 
                #   print("trying to open ",filename)
                with open(filename,'w') as _: pass
                tryagain = False
            except PermissionError: 
                print()
                t = input("cannot open {}".format(filename)+
                          "\nPermission denied, try again? [Y/n]")
                if t.lower() in ["no","n"]: 
                    tryagain = False
                    writefile = False
                
        if writefile is True: wb.save(filename=filename)
        if verbose is True: print("wrote {}".format(filename))
    return wb

def matml(materialdata,filename,verbose=False,ansys=True):
    """ writes MatML file [*]_
    
    .. [*] not yet completely compatible with ANSYS
    """
    
    from materialtools import Material
    from materialtools import MaterialProperty as Property
    from materialtools import MaterialParameter as Parameter

    def write_material(material,indent=1):
        """
        """
        ''' write the top lines'''
        material_lines = ["<Material>\n\t<BulkDetails>"]
        
        ''' get the non-property material atributes'''
        othervalues = [p for p in material if type(material[p]) is not Property]
        #for value in othervalues:     
        for value in ["Name","Description"]:     
            ''' write each non-property attribute to a line'''
            material_lines.append('\t\t<{0:}>{1:}</{0:}>'.format(value,material[value]))
        
        ''' get the list of properties for this material'''
        properties = [p for p in material.values() if type(p) is Property]
        for mproperty in properties:
            '''write the lines for each property '''
            material_lines.append(write_property(mproperty))
            
        ''' close the tags'''    
        material_lines.append("\t</BulkDetails>\n</Material>")
        linesep = "\n"+"\t"*indent
        return linesep.join(material_lines)
        
    def write_property(materialproperty,indent=2):
        """ write property
        
        runs through a material property and writes parameters
        """
        ''' find the id'''
        id = materialdata.ids[materialproperty["Name"]]

        ''' write the first line'''
        property_lines = ["\t\t<PropertyData property=\"{}\">".format(id)]
        
        ''' get the list of parameters'''
        parameters = [pa for pa in materialproperty.values() if type(pa) is Parameter]
        for parameter in parameters:
            
            ''' write the lines for each parameter'''
            property_lines.append(write_parameter(parameter))
            
        ''' write the closing tag on the last line'''    
        property_lines.append("</PropertyData>")
        linesep = "\n"+"\t"*indent
        return linesep.join(property_lines)

    def write_parameter(materialparameter,indent=3):
        """ write parameter
        
        runs through qualifiers in a parameter and writes them as xml
        """
        name = materialparameter["Name"]
        mid = materialdata.ids[name]
        
        data = materialparameter["Values"]

        dformat = str(type(data[0]).__name__)
        if dformat == "str": dformat = "string"
        elif dformat == "NoneType": dformat = ""
        elif dformat == "int": dformat = "float"
        
        try: data = ','.join([str(x) for x in data])
        except: pass
        if '-' in data: data = "-"
        qualifiers = []
        
        parameter_lines = ["<ParameterValue parameter='{}' format=\"{}\">".format(mid,dformat)]
        parameter_lines.append("<Data format=\"{}\">{}</Data>".format(dformat,data))
        
        for qualifier in qualifiers:
            parameter_lines.append(write_qualifier(qualifier))
            
        parameter_lines.append("</ParameterValue>")
        linesep = "\n"+"\t"*indent
        return linesep.join(parameter_lines)
                
    def write_qualifier(qualifier,indent=4):
        """
        """
        name = ''
        data = ''
        #name = qualifier.keys()
        #content = (x for x in qualifier.values())
        qualifier_lines = "<Qualifier name =\"{}\">{}</Qualifier>".format(name,data)
        
        linesep = "\n"+"\t"*indent
        return linesep.join(qualifier_lines)

    def write_metadata(materialdata,indent=2):
        mdata_lines = ["<Metadata>"]
        properties = [pr for pr in materialdata.ids.items() if "pr" in pr[1]]
        parameters = [pa for pa in materialdata.ids.items() if "pa" in pa[1]]        
        properties.sort(key=lambda x: x[1])
        parameters.sort(key=lambda x: x[1])
        for pa in parameters:
            name, paid = pa
            mdata_lines.append("<ParameterDetails id=\"{}\">".format(paid))
            mdata_lines.append("\t<Name>{}</Name>".format(name))
            mdata_lines.append(write_unitlines(materialdata,paid))
            mdata_lines.append("</ParameterDetails>")
        
        for pr in properties:
            name, prid = pr
            mdata_lines.append("<PropertyDetails id=\"{}\">".format(prid))
            mdata_lines.append("\t<Name>{}</Name>".format(name))
            mdata_lines.append(write_unitlines(materialdata,prid))
            mdata_lines.append("</PropertyDetails>")
        

        
        mdata_lines.append("</Metadata>")
        linesep = "\n"+"\t"*indent
        return linesep.join(mdata_lines)
    
    def write_unitlines(materialdata,pid,indent = 1):
        unitlines = ["\t<Unitless />"]
        linesep = "\n"+"\t"*indent
        return linesep.join(unitlines)

    def write_matml_lines(materialdata,indent = 1):
        try: materialdata.generate_ids()
        except: print("could not generate ids for metadata"); raise
        materials = [m for m in materialdata.values() if type(m) is Material]
        #print([m.name for m in materials])    
        matml_lines = ["<MatML_Doc>"]
        for material in materials:
            matml_lines.append(write_material(material))
            
        matml_lines.append(write_metadata(materialdata))
        matml_lines.append("</MatML_Doc>")
        
        linesep = "\n"+"\t"*indent
        return linesep.join(matml_lines)

    matml_lines = write_matml_lines(materialdata)
    if ansys is True:
        xml_wrapper = ["<?xml version=\"1.0\" encoding=\"UTF-8\"?>\n"+
                       "<EngineeringData version='15.0.0.504' versiondate='16/10/2013 16:34:00'>\n"+
                       "\t<Notes>\n"+
                       "\t</Notes>\n"+
                       "<Materials>\n",
                        "\n</Materials>\n"+
                        "</EngineeringData>"]
        matml_lines = xml_wrapper[0]+matml_lines+xml_wrapper[1]
        
    try:
        with open(filename,"w") as f:
            f.writelines(matml_lines)
    except:
        print("no file output")
        #raise
    
    return matml_lines

def list_contents(material):
    """ lists the contents of a material object
    
        **hasn't been used in a while so may be broken**
    
    """
    from materialtools import Material, MaterialProperty, MaterialParameter
    assert type(material) == Material, "argument must be a Material object"
    
    print('\n',material.name)
    print('='*80)
    properties = [x for x in material.values() if type(x) is MaterialProperty]
    for materialproperty in properties:
            print('\n',materialproperty.name)
            print('-'*80)
            parameters = [p for p in materialproperty.values() if type(p) is MaterialParameter]
            for parameter in parameters:
                    print(' {:} [{}]'.format(parameter.name,parameter["Units"][0]))
                    [print('{:}'.format(v),end=',') for v in parameter["Values"]]
    print('_'*80)
    return