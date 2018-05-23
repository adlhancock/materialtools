# -*- coding: utf-8 -*-
"""methods for importing material property files

Created on Wed Nov  9 11:47:47 2016

.. moduleauthor:: adlhancock
"""

        
def matml(filename = None, 
          materialname = 'auto',
          filetype = 'ANSYS', 
          testing = False, 
          verbose = False):
    
    """Uses xmlto dict to import MatML data in xml format
    
    returns:
        materialdata
    """

    from materialtools import MatMLData
    materialdata = MatMLData()
    
    if testing is True:
        filename = './sampledata/testdata.xml'
        filename = '/python/materialtools/sampledata/testdata.xml'
 
    def importxml(filename,filetype='ANSYS'):
        import xmltodict
        with open(filename,'r') as f:
            xml = f.read()
        data = xmltodict.parse(xml)
        if filetype == 'ANSYS':
            matml = data['EngineeringData']['Materials']['MatML_Doc']
        else: 
            print('Only ANSYS files handled at the moment')
            matml = {}
        return matml

    try:
        materialdata.matml = importxml(filename)
        if verbose is True: print('[read.py] matml file imported:',filename)
    except: print('ERROR: no data imported'); raise
    
    try:         
        materialdata.getdata(verbose)
    except:
        print('[read.py] ERROR: could not get data from matml file')
        raise
    
    #materialdata['Source'] = filename
    materialdata.filename = filename
    return materialdata
    
def xlsx(filename, 
         materialname = 'auto',
         filetype = 'xlsx',
         testing = False,
         verbose = False):

    """Imports an xlsx file 
    returns: 
        materialdata
    """
    
    from materialtools import (MaterialData, 
                               Material, 
                               MaterialProperty, 
                               MaterialParameter)
    from openpyxl import load_workbook
    
    material = Material()
    material.filename = filename
    
    # open workbook
    wb = load_workbook(filename)
    
    # get metadata information
    if verbose is True: 
        print(filename)
    #ws = wb.get_sheet_by_name("Metadata") ## deprecated
    ws = wb["Metadata"]
    for row in ws.rows:
        if row[0].value not in (
                "Propertynames",
                "Abbreviations",
                "DataSources",
                "Comments"):
            material[row[0].value] = row[1].value
    if materialname is 'auto':
        #print('auto')
        material.name = material["MaterialName"]
    else:
        material.name = materialname
        material["MaterialName"] = materialname
    
    
    ## list the filename as the source if none otherwise specified    
    try: 
        material.source = material["DataSource"]
    except: 
        material.source = material.filename
        material["DataSource"] = material.source

    # if material condition not listed, use data source
    if material["Condition"] in ['','Unknown',None]:
        material["Condition"] = material["DataSource"]

    # get dictionary of abbreviations
    propertynames,abbreviations = [],[]
    for row in ws.rows:
        if row[0].value == "Propertynames":
            propertynames = [cell.value for cell in row[1:]]
        if row[0].value == "Abbreviations":
            abbreviations = [cell.value for cell in row[1:]]
        if row[0].value == "DataSources":
            datasources = [cell.value for cell in row[1:]]
        if row[0].value == "Comments":
            comments = [cell.value for cell in row[1:]]
    assert abbreviations != [], "no property abbreviations found"
    propertydict = {x[0]:x[1] for x in zip(abbreviations,propertynames)}
    try:
        commentsdict = {x[0]:x[1] for x in zip(abbreviations,comments)}
    except:
        commentsdict = {x:"" for x in abbreviations}
            
    try:
        sourcedict = {x[0]:x[1] for x in zip(abbreviations,datasources)}
    except:
        sourcedict = {x:material.source for x in abbreviations}
    #print('\n\n\n',sourcedict)

    # get names of remaining worksheets
    #sheetnames = wb.get_sheet_names() ## deprecated
    sheetnames = wb.sheetnames
    sheetnames.pop(sheetnames.index("Metadata"))
    for sheetname in sheetnames:
        #ws = wb.get_sheet_by_name(sheetname) ## deprecated
        ws = wb[sheetname]
        propertyname = propertydict[sheetname]
        propertysource = sourcedict[sheetname]
        propertycomments = commentsdict[sheetname]
        materialproperty = MaterialProperty(name = propertyname, 
                                            source = propertysource,
                                            comments = propertycomments)
        #materialproperty["DataSource"] = propertysource

        # write parameters        
        parameternames = [cell.value for cell in [r for r in ws.rows][0]]
        parameterunits = [cell.value for cell in [r for r in ws.rows][1]]
        parametervalues = []
        for i,parametername in enumerate(parameternames):
            parametervalues.append(
                [cell.value for cell in [c for c in ws.columns][i]][2:])
        for i,p in enumerate(parameternames):
            #print(parameterunits[i]*len(parametervalues[i]))
            if parameterunits[i] is None: parameterunits[i] = ['']
            units = [parameterunits[i]]*len(parametervalues[i])
            materialproperty[p] = MaterialParameter(
                                    parameternames[i],
                                    units, 
                                    parametervalues[i])
        
        material[propertyname] = materialproperty

    material.propertynames = propertynames
    material.filename = filename
    material["Filename"] = filename
    materialdata = MaterialData()
    materialdata[material["MaterialName"]] = material
    if verbose is True: 
        print('Imported: {} from {}'.format(material["MaterialName"],material.source))
    return materialdata

def textfile(filename = None, 
             materialname = 'auto',
             filetype = 'text', 
             testing = False,
             verbose = False):
    '''
    Imports a text file and returns a nested dictionary of material properties.
    '''
    from materialtools import Material, MaterialData
    if testing is True:
        filename = "/python/materialtools/sampledata/testdata.txt"
        print('using test data from {}'.format(filename))

    with open(filename) as file:
        rawtext = file.read()
    lines = rawtext.splitlines()
    lines = [line for line in lines if line is not '']

    materialdata = MaterialData()

    firstheading = lines[0][0:lines[0].find('=')].strip()
    materialname = lines[0][lines[0].find('=')+1:].strip()

    if firstheading == 'Material Name':
        materialdata[materialname] = Material(materialname=materialname, 
                                              verbose=False)
        
    else:
        print("the first line must be in the form 'Material Name = [materialname]'")

    for line in lines:
        # get the variable title
        variable = line[0:line.find('=')].strip()
        # if this is not the start of a new material...
        if variable not in ['Material Name']:
            try: # try to get numerical data
                value = float(line[line.find('=')+1:line.find('[')].strip())
                units = line[line.find('[')+1:line.find(']')].strip()
                materialdata[materialname][variable] = {}
                materialdata[materialname][variable] = {
                                        'Values':value,'Units':units}
                materialdata[materialname][variable][variable] = {
                                        'Values':value,'Units':units}
            except: # if this doesn't work, just take the line after the '='
                materialdata[materialname][variable] = line[line.find('=')+1:].strip()
        else:
            materialname = line[line.find('=')+1:].strip()
            materialdata[materialname] = Material(materialname=materialname, 
                                                  verbose=False)

    if verbose is True:
        print('imported the followng materials from ',filename,":")
        print([materialname for materialname in materialdata])
        #materialdata.list_contents()
    #print(materialdata)
    
    return materialdata

def csv(filename,
        materialname = 'auto',
        filetype='csv',
        testing=False):
    """ imports a csv file
    """
    
    import csv
    from materialtools import Material, MaterialData
    material = Material()
    rawdata = []

    # import data from file
    with open(filename, mode = 'r') as importedfile:
        for row in csv.reader(importedfile):
            rawdata.append(row)

    materialnames = [row[0] for rows in [row for row in rawdata if "Name" in row[0]]]
    print(materialnames)
    # strip material name from top of file
    if "Name" in rawdata[0][0]:
        materialname = rawdata[0][1]
        materialdata = rawdata[1:]

    # strip property names from next line
    propertynames = [propertyname.strip() for propertyname in materialdata[0]]
    allvalues = materialdata[1:]

    # check left hand column is temperature and get temperature values
    assert propertynames[0] == "Temperature", "Temperature must be first column"
    temperaturevalues = [row[0] for row in allvalues]

    # populate material property values
    for n,propertyname in enumerate(propertynames):
        if propertyname not in ['','Temperature']:

            # set up data structure
            material[propertyname] = {propertyname:{"Units":'',"Values":[]},
                            "Temperature":{
                                "Units":'',"Values":[
                                    x for x in temperaturevalues]},
                            "Units":'',
                            "Values":[]}

            # fill in numbers
            for i, row in enumerate(allvalues):
                try:
                    # get the datapoints in float format
                    propval = float(row[n].strip())
                    material[propertyname]["Values"].append(propval)
                    material[
                        propertyname][propertyname]["Values"].append(propval)
                except:
                    # say if there's no datapoint at this temperature
                    missingt = material[
                        propertyname]["Temperature"]["Values"][i]
                    print("no value for",propertyname,"at",missingt)

                    # mask off the temperature point
                    material[propertyname]["Temperature"]["Values"][i]=None

            # clean up the empty rows
            material[propertyname]["Temperature"]["Values"] = [
                x for x in material[
                    propertyname]["Temperature"]["Values"] if x is not None]

    # put the material data for each material in materialdata
    materialdata = MaterialData()
    materialdata.update({materialname:material})
    print("imported {} from {}".format(materialname,filename))
    return materialdata
    



if __name__ == '__main__':
    import materialtools
    materialdata = materialtools.MaterialData()
    materialdata.import_file(testing=True)
    materialdata.export_file('F:/output.json')